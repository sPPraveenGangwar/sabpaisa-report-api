"""
QwikForms Views
Admin-only views for QwikForms transaction history, settlements, analytics, and reports
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Count, Sum, Avg, F, Case, When, DecimalField
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv
import io

from .models import DataTransactions, DataForm, DataFormDetails, CollegeMaster, ClientMappingCode, BankDetailsBean
from .serializers import (
    DataTransactionsSerializer,
    DataTransactionsDetailedSerializer,
    QwikFormsTransactionWithFormSerializer,
    QwikFormsSettlementSerializer,
    QwikFormsRefundSerializer,
    QwikFormsAnalyticsSerializer
)
from apps.core.permissions import IsAdmin


class QwikFormsTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Admin-only QwikForms Transaction History

    GET /api/v1/qwikforms/transactions/ - List all transactions with form data
    GET /api/v1/qwikforms/transactions/{id}/ - Get single transaction detail
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    serializer_class = QwikFormsTransactionWithFormSerializer

    def get_queryset(self):
        """Get transactions from QwikForms database with filters - Optimized for high volume"""
        from django.db.models import F as DjangoF, Value, CharField
        from django.db.models.functions import Coalesce

        # Use annotate to fetch client_name and client_code from college_master via JOIN
        queryset = DataTransactions.objects.using('qwikforms_db').annotate(
            client_name=Coalesce(
                DjangoF('college_id_fk__college_name'),
                Value(''),
                output_field=CharField()
            ),
            client_code=Coalesce(
                DjangoF('college_id_fk__college_code'),
                Value(''),
                output_field=CharField()
            )
        ).all()

        # Date filters - CRITICAL: Always use indexed date fields
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if date_from:
            queryset = queryset.filter(trans_date__gte=date_from)
        if date_to:
            # Add one day to include the end date
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(trans_date__lt=date_to_obj)

        # Transaction status filter
        trans_status = self.request.query_params.get('trans_status')
        if trans_status:
            queryset = queryset.filter(trans_status__iexact=trans_status)

        # Payment mode filter
        trans_paymode = self.request.query_params.get('trans_paymode')
        if trans_paymode:
            queryset = queryset.filter(trans_paymode__iexact=trans_paymode)

        # Client filter by client_code (more efficient than client_id for joins)
        client_code = self.request.query_params.get('client_code')
        if client_code:
            queryset = queryset.filter(client_code=client_code)

        # Form filter
        form_id = self.request.query_params.get('form_id')
        if form_id:
            queryset = queryset.filter(form_id=form_id)

        # Settlement status filter
        settlement_status = self.request.query_params.get('settlement_status')
        if settlement_status:
            queryset = queryset.filter(settlement_status__iexact=settlement_status)

        # Search filter - Use indexed fields only
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(trans_id__icontains=search) |
                Q(sp_trans_id__icontains=search)
            )

        return queryset.order_by('-trans_date')

    def list(self, request, *args, **kwargs):
        """List transactions with pagination"""
        queryset = self.filter_queryset(self.get_queryset())

        # Pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'success': True,
            'count': queryset.count(),
            'results': serializer.data
        })

    def retrieve(self, request, *args, **kwargs):
        """Get single transaction with complete form details"""
        instance = self.get_object()
        serializer = DataTransactionsDetailedSerializer(instance)
        return Response({
            'success': True,
            'data': serializer.data
        })

    @action(detail=False, methods=['get'], url_path='clients-and-forms')
    def clients_and_forms(self, request):
        """
        Get all clients from college_master and their related forms from data_form_details

        GET /api/v1/qwikforms/transactions/clients-and-forms/

        Returns:
        {
            "success": true,
            "data": {
                "clients": [
                    {
                        "college_id": 1,
                        "college_code": "ABC001",
                        "college_name": "ABC College",
                        "forms": [
                            {
                                "id": 1,
                                "form_name": "Application Form 2024",
                                "form_start_date": "2024-01-01",
                                "form_end_date": "2024-12-31"
                            }
                        ]
                    }
                ]
            }
        }
        """
        try:
            # Fetch all clients from college_master
            clients = CollegeMaster.objects.using('qwikforms_db').all().values(
                'college_id',
                'college_code',
                'college_name',
                'email_id',
                'contact'
            ).order_by('college_name')

            # Fetch all forms with their owner details (no status filter for reporting)
            forms = DataFormDetails.objects.using('qwikforms_db').all().values(
                'id',
                'form_name',
                'form_owner_id',
                'form_owner_name',
                'form_start_date',
                'form_end_date',
                'status'
            ).order_by('form_name')

            # Create a mapping of client_id -> forms
            client_form_map = {}
            for form in forms:
                owner_id = form['form_owner_id']
                if owner_id not in client_form_map:
                    client_form_map[owner_id] = []
                client_form_map[owner_id].append({
                    'id': form['id'],
                    'form_name': form['form_name'],
                    'form_start_date': form['form_start_date'].strftime('%Y-%m-%d') if form['form_start_date'] else None,
                    'form_end_date': form['form_end_date'].strftime('%Y-%m-%d') if form['form_end_date'] else None,
                    'status': form['status']
                })

            # Build response with clients and their forms
            clients_with_forms = []
            for client in clients:
                client_data = {
                    'college_id': client['college_id'],
                    'college_code': client['college_code'],
                    'college_name': client['college_name'],
                    'email': client['email_id'],
                    'contact': client['contact'],
                    'forms': client_form_map.get(client['college_id'], [])
                }
                clients_with_forms.append(client_data)

            return Response({
                'success': True,
                'data': {
                    'clients': clients_with_forms,
                    'total_clients': len(clients_with_forms),
                    'total_forms': len(forms)
                }
            })

        except Exception as e:
            return Response({
                'success': False,
                'message': f'Error fetching clients and forms: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class QwikFormsSettlementViewSet(viewsets.ViewSet):
    """
    Admin-only QwikForms Settlement APIs

    GET /api/v1/qwikforms/settlements/settled/ - Settled transactions
    GET /api/v1/qwikforms/settlements/pending/ - Pending settlements
    GET /api/v1/qwikforms/settlements/refunds/ - Refund transactions
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def _get_base_queryset(self):
        """Get base queryset with common filters"""
        queryset = DataTransactions.objects.using('qwikforms_db').all()

        # Date filters
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if date_from:
            queryset = queryset.filter(trans_date__gte=date_from)
        if date_to:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(trans_date__lt=date_to_obj)

        # Client filter
        client_id = self.request.query_params.get('client_id')
        if client_id:
            queryset = queryset.filter(client_id=client_id)

        return queryset

    @action(detail=False, methods=['get'])
    def settled(self, request):
        """Get settled transactions"""
        queryset = self._get_base_queryset().filter(
            trans_status='SUCCESS',
            settlement_status='COMPLETED',
            is_settled='Y'
        ).order_by('-settlement_date')

        serializer = QwikFormsSettlementSerializer(queryset, many=True)

        # Calculate totals
        totals = queryset.aggregate(
            total_trans_amount=Sum('trans_amount'),
            total_settlement_amount=Sum('settlement_amount'),
            total_charges=Sum('trans_charges'),
            count=Count('id')
        )

        return Response({
            'success': True,
            'count': totals['count'],
            'summary': {
                'total_transaction_amount': totals['total_trans_amount'] or 0,
                'total_settlement_amount': totals['total_settlement_amount'] or 0,
                'total_charges': totals['total_charges'] or 0
            },
            'results': serializer.data
        })

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get pending settlement transactions"""
        queryset = self._get_base_queryset().filter(
            trans_status='SUCCESS'
        ).filter(
            Q(settlement_status__in=['PENDING', 'PROCESSING']) | Q(is_settled='N')
        ).order_by('-trans_date')

        serializer = QwikFormsSettlementSerializer(queryset, many=True)

        totals = queryset.aggregate(
            total_amount=Sum('trans_amount'),
            count=Count('id')
        )

        return Response({
            'success': True,
            'count': totals['count'],
            'summary': {
                'total_pending_amount': totals['total_amount'] or 0
            },
            'results': serializer.data
        })

    @action(detail=False, methods=['get'])
    def refunds(self, request):
        """Get refund transactions"""
        queryset = self._get_base_queryset().filter(
            refund_id__isnull=False
        ).exclude(refund_id='').order_by('-trans_date')

        serializer = QwikFormsRefundSerializer(queryset, many=True)

        # Note: refund_amount is CharField, need to convert for aggregation
        results = serializer.data
        total_refund = sum(
            Decimal(item['refund_amount'])
            for item in results
            if item.get('refund_amount') and item['refund_amount'].strip()
        )

        return Response({
            'success': True,
            'count': len(results),
            'summary': {
                'total_refund_amount': total_refund
            },
            'results': results
        })


class QwikFormsAnalyticsViewSet(viewsets.ViewSet):
    """
    Admin-only QwikForms Analytics APIs

    GET /api/v1/qwikforms/analytics/dashboard/ - Complete analytics dashboard
    GET /api/v1/qwikforms/analytics/by-form/ - Form-wise analytics
    GET /api/v1/qwikforms/analytics/by-client/ - Client-wise analytics
    GET /api/v1/qwikforms/analytics/payment-mode/ - Payment mode distribution
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def _get_base_queryset(self):
        """Get base queryset with date filters"""
        queryset = DataTransactions.objects.using('qwikforms_db').all()

        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if date_from:
            queryset = queryset.filter(trans_date__gte=date_from)
        if date_to:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(trans_date__lt=date_to_obj)

        return queryset

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get complete analytics dashboard"""
        queryset = self._get_base_queryset()

        # Summary statistics
        summary = queryset.aggregate(
            total_transactions=Count('id'),
            successful_transactions=Count('id', filter=Q(trans_status='SUCCESS')),
            failed_transactions=Count('id', filter=Q(trans_status='FAILED')),
            pending_transactions=Count('id', filter=Q(trans_status='PENDING')),
            total_volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS')),
            total_settled_amount=Sum('settlement_amount', filter=Q(is_settled='Y')),
            pending_settlement_amount=Sum('trans_amount', filter=Q(trans_status='SUCCESS', is_settled='N')),
            avg_transaction_value=Avg('trans_amount', filter=Q(trans_status='SUCCESS'))
        )

        # Calculate success rate
        total = summary['total_transactions'] or 1
        summary['success_rate'] = (summary['successful_transactions'] / total) * 100 if total > 0 else 0

        # Payment mode distribution
        by_payment_mode = {}
        payment_modes = queryset.values('trans_paymode').annotate(
            count=Count('id'),
            volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS')),
            success_count=Count('id', filter=Q(trans_status='SUCCESS'))
        )

        for mode in payment_modes:
            mode_name = mode['trans_paymode'] or 'UNKNOWN'
            by_payment_mode[mode_name] = {
                'count': mode['count'],
                'volume': float(mode['volume'] or 0),
                'success_rate': (mode['success_count'] / mode['count']) * 100 if mode['count'] > 0 else 0
            }

        # Form-wise analytics
        by_form = queryset.filter(trans_status='SUCCESS').values('form_id', 'fee_name').annotate(
            count=Count('id'),
            volume=Sum('trans_amount')
        ).order_by('-volume')[:10]

        # Client-wise analytics
        by_client = queryset.filter(trans_status='SUCCESS').values('client_id').annotate(
            count=Count('id'),
            volume=Sum('trans_amount')
        ).order_by('-volume')[:10]

        # Add client names
        for client in by_client:
            if client['client_id']:
                try:
                    college = CollegeMaster.objects.using('qwikforms_db').filter(
                        college_code=client['client_id']
                    ).first()
                    client['client_name'] = college.college_name if college else client['client_id']
                except Exception:
                    client['client_name'] = client['client_id']

        # Daily trend (last 30 days)
        daily_trend = queryset.annotate(
            date=TruncDate('trans_date')
        ).values('date').annotate(
            count=Count('id'),
            volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS')),
            success_count=Count('id', filter=Q(trans_status='SUCCESS'))
        ).order_by('date')

        for day in daily_trend:
            day['success_rate'] = (day['success_count'] / day['count']) * 100 if day['count'] > 0 else 0

        return Response({
            'success': True,
            'data': {
                'summary': summary,
                'by_payment_mode': by_payment_mode,
                'by_form': list(by_form),
                'by_client': list(by_client),
                'daily_trend': list(daily_trend)
            }
        })

    @action(detail=False, methods=['get'], url_path='by-form')
    def by_form(self, request):
        """Get form-wise analytics"""
        queryset = self._get_base_queryset()

        analytics = queryset.values('form_id', 'fee_name').annotate(
            total_count=Count('id'),
            success_count=Count('id', filter=Q(trans_status='SUCCESS')),
            failed_count=Count('id', filter=Q(trans_status='FAILED')),
            total_volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS')),
            avg_amount=Avg('trans_amount', filter=Q(trans_status='SUCCESS'))
        ).order_by('-total_volume')

        # Calculate success rate for each form
        results = []
        for form in analytics:
            total = form['total_count'] or 1
            results.append({
                'form_id': form['form_id'],
                'form_name': form['fee_name'],
                'total_transactions': form['total_count'],
                'successful_transactions': form['success_count'],
                'failed_transactions': form['failed_count'],
                'success_rate': (form['success_count'] / total) * 100,
                'total_volume': float(form['total_volume'] or 0),
                'average_amount': float(form['avg_amount'] or 0)
            })

        return Response({
            'success': True,
            'count': len(results),
            'results': results
        })

    @action(detail=False, methods=['get'], url_path='by-client')
    def by_client(self, request):
        """Get client-wise analytics"""
        queryset = self._get_base_queryset()

        analytics = queryset.values('client_id', 'college_id_fk').annotate(
            total_count=Count('id'),
            success_count=Count('id', filter=Q(trans_status='SUCCESS')),
            total_volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS')),
            settled_amount=Sum('settlement_amount', filter=Q(is_settled='Y'))
        ).order_by('-total_volume')

        # Add client names
        results = []
        for client in analytics:
            client_name = client['client_id']
            if client['college_id_fk']:
                try:
                    college = CollegeMaster.objects.using('qwikforms_db').get(
                        college_id=client['college_id_fk']
                    )
                    client_name = college.college_name
                except CollegeMaster.DoesNotExist:
                    pass

            results.append({
                'client_id': client['client_id'],
                'client_name': client_name,
                'total_transactions': client['total_count'],
                'successful_transactions': client['success_count'],
                'total_volume': float(client['total_volume'] or 0),
                'settled_amount': float(client['settled_amount'] or 0)
            })

        return Response({
            'success': True,
            'count': len(results),
            'results': results
        })

    @action(detail=False, methods=['get'], url_path='payment-mode')
    def payment_mode(self, request):
        """Get payment mode distribution"""
        queryset = self._get_base_queryset()

        analytics = queryset.values('trans_paymode').annotate(
            total_count=Count('id'),
            success_count=Count('id', filter=Q(trans_status='SUCCESS')),
            failed_count=Count('id', filter=Q(trans_status='FAILED')),
            total_volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS')),
            avg_amount=Avg('trans_amount', filter=Q(trans_status='SUCCESS'))
        ).order_by('-total_volume')

        results = []
        for mode in analytics:
            total = mode['total_count'] or 1
            results.append({
                'payment_mode': mode['trans_paymode'] or 'UNKNOWN',
                'total_transactions': mode['total_count'],
                'successful_transactions': mode['success_count'],
                'failed_transactions': mode['failed_count'],
                'success_rate': (mode['success_count'] / total) * 100,
                'total_volume': float(mode['total_volume'] or 0),
                'average_amount': float(mode['avg_amount'] or 0)
            })

        return Response({
            'success': True,
            'results': results
        })


class QwikFormsReportViewSet(viewsets.ViewSet):
    """
    Admin-only QwikForms Report Generation

    POST /api/v1/qwikforms/reports/generate-excel/ - Generate Excel report
    POST /api/v1/qwikforms/reports/generate-csv/ - Generate CSV report
    POST /api/v1/qwikforms/reports/generate-pdf/ - Generate PDF report (HTML)
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def _get_report_data(self, request):
        """Get filtered data for reports"""
        queryset = DataTransactions.objects.using('qwikforms_db').all()

        # Apply filters
        filters = request.data.get('filters', {})

        if filters.get('date_from'):
            queryset = queryset.filter(trans_date__gte=filters['date_from'])
        if filters.get('date_to'):
            date_to_obj = datetime.strptime(filters['date_to'], '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            queryset = queryset.filter(trans_date__lt=date_to_obj)

        if filters.get('trans_status'):
            queryset = queryset.filter(trans_status__iexact=filters['trans_status'])

        if filters.get('trans_paymode'):
            queryset = queryset.filter(trans_paymode__iexact=filters['trans_paymode'])

        if filters.get('client_id'):
            queryset = queryset.filter(client_id=filters['client_id'])

        if filters.get('form_id'):
            queryset = queryset.filter(form_id=filters['form_id'])

        return queryset.order_by('-trans_date')

    @action(detail=False, methods=['post'], url_path='generate-excel')
    def generate_excel(self, request):
        """Generate Excel report with transaction and form data"""
        queryset = self._get_report_data(request)

        # Create workbook
        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Add summary statistics
        summary_stats = queryset.aggregate(
            total_count=Count('id'),
            success_count=Count('id', filter=Q(trans_status='SUCCESS')),
            failed_count=Count('id', filter=Q(trans_status='FAILED')),
            total_volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS')),
            settled_amount=Sum('settlement_amount', filter=Q(is_settled='Y'))
        )

        ws_summary['A1'] = 'QwikForms Transaction Report'
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A3'] = 'Total Transactions:'
        ws_summary['B3'] = summary_stats['total_count']
        ws_summary['A4'] = 'Successful Transactions:'
        ws_summary['B4'] = summary_stats['success_count']
        ws_summary['A5'] = 'Failed Transactions:'
        ws_summary['B5'] = summary_stats['failed_count']
        ws_summary['A6'] = 'Total Volume (INR):'
        ws_summary['B6'] = float(summary_stats['total_volume'] or 0)
        ws_summary['A7'] = 'Settled Amount (INR):'
        ws_summary['B7'] = float(summary_stats['settled_amount'] or 0)

        # Transactions sheet
        ws_trans = wb.create_sheet("Transactions")

        # Headers
        headers = [
            'Transaction ID', 'SP Trans ID', 'Date', 'Amount', 'Status', 'Payment Mode',
            'Customer Name', 'Email', 'Contact', 'Fee Name', 'Client ID',
            'PG Trans ID', 'Bank Ref No', 'Settlement Status', 'Settlement Date',
            'Settlement Amount', 'Charges'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)

        # Add data
        for row, trans in enumerate(queryset[:5000], 2):  # Limit to 5000 rows
            ws_trans.cell(row=row, column=1, value=trans.trans_id)
            ws_trans.cell(row=row, column=2, value=trans.sp_trans_id)
            ws_trans.cell(row=row, column=3, value=trans.trans_date.strftime('%Y-%m-%d %H:%M:%S') if trans.trans_date else '')
            ws_trans.cell(row=row, column=4, value=float(trans.trans_amount) if trans.trans_amount else 0)
            ws_trans.cell(row=row, column=5, value=trans.trans_status)
            ws_trans.cell(row=row, column=6, value=trans.trans_paymode)
            ws_trans.cell(row=row, column=7, value=trans.name)
            ws_trans.cell(row=row, column=8, value=trans.email)
            ws_trans.cell(row=row, column=9, value=trans.contact)
            ws_trans.cell(row=row, column=10, value=trans.fee_name)
            ws_trans.cell(row=row, column=11, value=trans.client_id)
            ws_trans.cell(row=row, column=12, value=trans.pg_trans_id)
            ws_trans.cell(row=row, column=13, value=trans.bank_reference_no)
            ws_trans.cell(row=row, column=14, value=trans.settlement_status)
            ws_trans.cell(row=row, column=15, value=trans.settlement_date.strftime('%Y-%m-%d') if trans.settlement_date else '')
            ws_trans.cell(row=row, column=16, value=float(trans.settlement_amount) if trans.settlement_amount else 0)
            ws_trans.cell(row=row, column=17, value=float(trans.trans_charges) if trans.trans_charges else 0)

        # Form Data sheet
        ws_forms = wb.create_sheet("Form Data")
        form_headers = ['Transaction ID', 'Form ID', 'Applicant Name', 'Email', 'Contact', 'Form Status', 'UDF Field 1', 'UDF Field 2']

        for col, header in enumerate(form_headers, 1):
            cell = ws_forms.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)

        form_row = 2
        for trans in queryset[:5000]:
            if trans.form_id:
                try:
                    form = DataForm.objects.using('qwikforms_db').get(form_id=trans.form_id)
                    ws_forms.cell(row=form_row, column=1, value=trans.trans_id)
                    ws_forms.cell(row=form_row, column=2, value=form.form_id)
                    ws_forms.cell(row=form_row, column=3, value=form.name)
                    ws_forms.cell(row=form_row, column=4, value=form.email)
                    ws_forms.cell(row=form_row, column=5, value=form.contact)
                    ws_forms.cell(row=form_row, column=6, value=form.form_status)
                    ws_forms.cell(row=form_row, column=7, value=form.udf_field1)
                    ws_forms.cell(row=form_row, column=8, value=form.udf_field2)
                    form_row += 1
                except DataForm.DoesNotExist:
                    pass

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"qwikforms_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['post'], url_path='generate-csv')
    def generate_csv(self, request):
        """Generate CSV report"""
        queryset = self._get_report_data(request)

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        headers = [
            'Transaction ID', 'SP Trans ID', 'Date', 'Amount', 'Status', 'Payment Mode',
            'Customer Name', 'Email', 'Contact', 'Fee Name', 'Client ID',
            'Form ID', 'Form Applicant', 'Form Email', 'Form Contact', 'Form Status',
            'Settlement Status', 'Settlement Date', 'Settlement Amount'
        ]
        writer.writerow(headers)

        # Write data
        for trans in queryset[:10000]:  # Limit to 10000 rows for CSV
            # Get form data if exists
            form_applicant = form_email = form_contact = form_status = ''
            if trans.form_id:
                try:
                    form = DataForm.objects.using('qwikforms_db').get(form_id=trans.form_id)
                    form_applicant = form.name or ''
                    form_email = form.email or ''
                    form_contact = form.contact or ''
                    form_status = form.form_status or ''
                except DataForm.DoesNotExist:
                    pass

            writer.writerow([
                trans.trans_id or '',
                trans.sp_trans_id or '',
                trans.trans_date.strftime('%Y-%m-%d %H:%M:%S') if trans.trans_date else '',
                float(trans.trans_amount) if trans.trans_amount else 0,
                trans.trans_status or '',
                trans.trans_paymode or '',
                trans.name or '',
                trans.email or '',
                trans.contact or '',
                trans.fee_name or '',
                trans.client_id or '',
                trans.form_id or '',
                form_applicant,
                form_email,
                form_contact,
                form_status,
                trans.settlement_status or '',
                trans.settlement_date.strftime('%Y-%m-%d') if trans.settlement_date else '',
                float(trans.settlement_amount) if trans.settlement_amount else 0
            ])

        # Create response
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        filename = f"qwikforms_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Add UTF-8 BOM for Excel compatibility
        response.write('\ufeff')

        return response

    @action(detail=False, methods=['post'], url_path='generate-pdf')
    def generate_pdf(self, request):
        """Generate HTML report for PDF printing"""
        queryset = self._get_report_data(request)

        # Summary statistics
        summary_stats = queryset.aggregate(
            total_count=Count('id'),
            success_count=Count('id', filter=Q(trans_status='SUCCESS')),
            total_volume=Sum('trans_amount', filter=Q(trans_status='SUCCESS'))
        )

        # Generate HTML
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>QwikForms Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #1976d2; border-bottom: 3px solid #1976d2; padding-bottom: 10px; }}
                .summary {{ margin: 20px 0; padding: 15px; background: #f5f5f5; border-radius: 5px; }}
                .summary-item {{ margin: 10px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 12px; }}
                th {{ background: #1976d2; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background: #f9f9f9; }}
                .footer {{ margin-top: 30px; text-align: center; color: #666; font-size: 11px; }}
            </style>
        </head>
        <body>
            <h1>QwikForms Transaction Report</h1>
            <div class="summary">
                <div class="summary-item"><strong>Total Transactions:</strong> {summary_stats['total_count']}</div>
                <div class="summary-item"><strong>Successful Transactions:</strong> {summary_stats['success_count']}</div>
                <div class="summary-item"><strong>Total Volume:</strong> ₹{summary_stats['total_volume'] or 0:,.2f}</div>
                <div class="summary-item"><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
            </div>

            <table>
                <thead>
                    <tr>
                        <th>Transaction ID</th>
                        <th>Date</th>
                        <th>Amount</th>
                        <th>Status</th>
                        <th>Mode</th>
                        <th>Customer</th>
                        <th>Fee Name</th>
                    </tr>
                </thead>
                <tbody>
        """

        # Add transaction rows (limit to 1000 for PDF)
        for trans in queryset[:1000]:
            html += f"""
                    <tr>
                        <td>{trans.trans_id or ''}</td>
                        <td>{trans.trans_date.strftime('%Y-%m-%d %H:%M') if trans.trans_date else ''}</td>
                        <td>₹{trans.trans_amount or 0:,.2f}</td>
                        <td>{trans.trans_status or ''}</td>
                        <td>{trans.trans_paymode or ''}</td>
                        <td>{trans.name or ''}</td>
                        <td>{trans.fee_name or ''}</td>
                    </tr>
            """

        html += """
                </tbody>
            </table>
            <div class="footer">
                <p>QwikForms Report - Generated by SabPaisa Reports System</p>
            </div>
        </body>
        </html>
        """

        # Create response
        response = HttpResponse(html, content_type='text/html; charset=utf-8')
        filename = f"qwikforms_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response