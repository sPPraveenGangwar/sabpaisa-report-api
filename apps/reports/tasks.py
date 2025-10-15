"""
Celery tasks for report generation with actual implementation
For development, these run synchronously. In production, remove the .delay assignments.
"""
import logging
from datetime import datetime
from typing import Dict, Any
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.utils import timezone

logger = logging.getLogger(__name__)


class MockTask:
    """Mock task object for development"""
    def __init__(self):
        self.id = f"task_{datetime.now().strftime('%Y%m%d%H%M%S')}"


def generate_transaction_excel(user_id: int, filters: Dict[str, Any], report_type: str = 'merchant') -> MockTask:
    """
    Generate transaction Excel report with comprehensive filters
    """
    logger.info(f"Generating {report_type} transaction Excel for user {user_id}")

    try:
        from apps.transactions.models import TransactionDetail
        from apps.transactions.filters import TransactionSearchFilter
        from apps.authentication.models import User

        # Get user for filtering
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            logger.error(f"User {user_id} not found for transaction excel")
            return MockTask()

        # Build queryset
        queryset = TransactionDetail.objects.all()

        # Apply comprehensive filters using TransactionSearchFilter
        queryset = TransactionSearchFilter.apply_filters(
            queryset,
            filters,
            user
        )

        # Limit records
        max_records = filters.get('max_records', 25000)
        queryset = queryset.order_by('-trans_date')[:max_records]

        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Headers
        headers = [
            'Transaction ID', 'Client Code', 'Client Name', 'Transaction Date',
            'Status', 'Payment Mode', 'Amount', 'Payee Email', 'Payee Mobile'
        ]

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Add data
        row_num = 2
        for txn in queryset:
            ws.cell(row=row_num, column=1, value=txn.txn_id)
            ws.cell(row=row_num, column=2, value=txn.client_code)
            ws.cell(row=row_num, column=3, value=txn.client_name)
            ws.cell(row=row_num, column=4, value=str(txn.trans_date) if txn.trans_date else '')
            ws.cell(row=row_num, column=5, value=txn.status)
            ws.cell(row=row_num, column=6, value=txn.payment_mode)
            ws.cell(row=row_num, column=7, value=txn.paid_amount)
            ws.cell(row=row_num, column=8, value=txn.payee_email)
            ws.cell(row=row_num, column=9, value=txn.payee_mob)
            row_num += 1

        # Save file
        os.makedirs('/tmp/reports', exist_ok=True)
        filename = f"transactions_{user_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('/tmp/reports', filename)
        wb.save(filepath)

        logger.info(f"Excel generated successfully: {filepath}")

    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")

    return MockTask()


generate_transaction_excel.delay = generate_transaction_excel  # Mock Celery delay method


def generate_settlement_excel(user_id: int, filters: Dict[str, Any]) -> MockTask:
    """
    Generate settlement Excel report with comprehensive filters
    """
    logger.info(f"Generating settlement Excel for user {user_id}")

    try:
        from apps.transactions.models import TransactionDetail
        from apps.transactions.filters import TransactionSearchFilter
        from apps.authentication.models import User

        # Get user for filtering
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            logger.error(f"User {user_id} not found for settlement excel")
            return MockTask()

        # Start with settled transactions
        queryset = TransactionDetail.objects.filter(is_settled=True)

        # Apply comprehensive filters
        queryset = TransactionSearchFilter.apply_filters(
            queryset,
            filters,
            user
        )

        # Apply settlement-specific date filters if specified
        use_settlement_date = filters.get('use_settlement_date', 'false').lower() == 'true'
        if use_settlement_date:
            date_filter = filters.get('date_filter', 'custom')
            if date_filter != 'custom':
                date_from, date_to = TransactionSearchFilter.get_date_range(date_filter)
                if date_from and date_to:
                    queryset = queryset.filter(settlement_date__range=[date_from, date_to])
            else:
                if filters.get('date_from'):
                    queryset = queryset.filter(settlement_date__gte=filters['date_from'])
                if filters.get('date_to'):
                    queryset = queryset.filter(settlement_date__lte=filters['date_to'])

        # Create Excel similar to transaction excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settlements"

        headers = [
            'Transaction ID', 'Settlement Date', 'Settlement Amount',
            'Settlement Status', 'Client Code', 'Payment Mode'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        row_num = 2
        for txn in queryset[:15000]:
            ws.cell(row=row_num, column=1, value=txn.txn_id)
            ws.cell(row=row_num, column=2, value=str(txn.settlement_date) if txn.settlement_date else '')
            ws.cell(row=row_num, column=3, value=txn.settlement_amount)
            ws.cell(row=row_num, column=4, value=txn.settlement_status)
            ws.cell(row=row_num, column=5, value=txn.client_code)
            ws.cell(row=row_num, column=6, value=txn.payment_mode)
            row_num += 1

        os.makedirs('/tmp/reports', exist_ok=True)
        filename = f"settlements_{user_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('/tmp/reports', filename)
        wb.save(filepath)

        logger.info(f"Settlement Excel generated: {filepath}")

    except Exception as e:
        logger.error(f"Error generating settlement Excel: {str(e)}")

    return MockTask()


generate_settlement_excel.delay = generate_settlement_excel


def generate_settlement_excel_v2(user_id: int, filters: Dict[str, Any], include_charts: bool = False) -> MockTask:
    """
    Generate enhanced settlement Excel report with charts and comprehensive filters
    """
    logger.info(f"Generating enhanced settlement Excel for user {user_id}")
    logger.info(f"Filters: {filters}, Include charts: {include_charts}")

    try:
        from apps.transactions.models import TransactionDetail
        from apps.transactions.filters import TransactionSearchFilter
        from apps.authentication.models import User
        from openpyxl.chart import BarChart, Reference, PieChart

        # Get user for filtering
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            logger.error(f"User {user_id} not found for settlement excel v2")
            return MockTask()

        # Start with settled transactions
        queryset = TransactionDetail.objects.filter(is_settled=True)

        # Apply comprehensive filters
        queryset = TransactionSearchFilter.apply_filters(
            queryset,
            filters,
            user
        )

        # Create Excel with enhanced features
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settlements"

        # Headers
        headers = [
            'Transaction ID', 'Client Code', 'Client Name', 'Settlement Date',
            'Settlement Amount', 'Settlement Status', 'Settlement UTR',
            'Payment Mode', 'Original Amount', 'Conv Charges', 'EP Charges', 'GST', 'Effective Amount'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Add data
        row_num = 2
        max_records = filters.get('max_records', 15000)
        for txn in queryset[:max_records]:
            ws.cell(row=row_num, column=1, value=txn.txn_id)
            ws.cell(row=row_num, column=2, value=txn.client_code)
            ws.cell(row=row_num, column=3, value=txn.client_name)
            ws.cell(row=row_num, column=4, value=str(txn.settlement_date) if txn.settlement_date else '')
            ws.cell(row=row_num, column=5, value=txn.settlement_amount)
            ws.cell(row=row_num, column=6, value=txn.settlement_status)
            ws.cell(row=row_num, column=7, value=txn.settlement_utr)
            ws.cell(row=row_num, column=8, value=txn.payment_mode)
            ws.cell(row=row_num, column=9, value=txn.paid_amount)
            ws.cell(row=row_num, column=10, value=txn.convcharges)
            ws.cell(row=row_num, column=11, value=txn.ep_charges)
            ws.cell(row=row_num, column=12, value=txn.gst)
            ws.cell(row=row_num, column=13, value=txn.effective_settlement_amount)
            row_num += 1

        # Add summary sheet if requested
        if include_charts:
            summary_ws = wb.create_sheet(title="Summary")
            summary_ws.append(['Settlement Summary'])
            summary_ws.append(['Total Records', queryset.count()])
            summary_ws.append(['Total Settlement Amount', queryset.aggregate(models.Sum('settlement_amount'))['settlement_amount__sum'] or 0])

        # Save file
        os.makedirs('/tmp/reports', exist_ok=True)
        filename = f"settlements_enhanced_{user_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('/tmp/reports', filename)
        wb.save(filepath)

        logger.info(f"Enhanced Settlement Excel generated: {filepath}")

    except Exception as e:
        logger.error(f"Error generating enhanced settlement Excel: {str(e)}")

    return MockTask()


generate_settlement_excel_v2.delay = generate_settlement_excel_v2


def generate_reconciliation_report(
    user_id: int,
    date_from: datetime,
    date_to: datetime,
    reconciliation_type: str = 'three_way'
) -> MockTask:
    """
    Generate reconciliation report
    """
    logger.info(f"Generating {reconciliation_type} reconciliation report for user {user_id}")
    logger.info(f"Date range: {date_from} to {date_to}")

    return MockTask()


generate_reconciliation_report.delay = generate_reconciliation_report


def export_analytics_report(
    user_id: int,
    report_type: str,
    filters: Dict[str, Any],
    format: str = 'excel'
) -> MockTask:
    """
    Export analytics report in various formats with comprehensive filters
    """
    logger.info(f"Exporting {report_type} analytics report for user {user_id}")
    logger.info(f"Format: {format}, Filters: {filters}")

    try:
        from apps.transactions.models import TransactionDetail
        from apps.transactions.filters import TransactionSearchFilter
        from apps.authentication.models import User
        from django.db import models

        # Get user for filtering
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            logger.error(f"User {user_id} not found for analytics export")
            return MockTask()

        # Build queryset
        queryset = TransactionDetail.objects.all()

        # Apply comprehensive filters
        queryset = TransactionSearchFilter.apply_filters(
            queryset,
            filters,
            user
        )

        if format == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Analytics_{report_type}"

            # Different headers based on report type
            if report_type == 'payment_mode':
                headers = [
                    'Payment Mode', 'Total Transactions', 'Successful', 'Failed',
                    'Success Rate', 'Total Volume', 'Average Transaction'
                ]
            elif report_type == 'merchant':
                headers = [
                    'Merchant Code', 'Merchant Name', 'Total Transactions',
                    'Success Rate', 'Total Volume', 'Settlement Pending'
                ]
            else:
                headers = [
                    'Date', 'Transactions', 'Successful', 'Failed',
                    'Volume', 'Settlement Amount'
                ]

            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Add aggregated data based on report type
            row_num = 2
            if report_type == 'payment_mode':
                payment_modes = queryset.values('payment_mode').annotate(
                    total=models.Count('txn_id'),
                    successful=models.Count('txn_id', filter=models.Q(status='SUCCESS')),
                    failed=models.Count('txn_id', filter=models.Q(status='FAILED')),
                    volume=models.Sum('paid_amount', filter=models.Q(status='SUCCESS')),
                    avg_amount=models.Avg('paid_amount', filter=models.Q(status='SUCCESS'))
                ).order_by('-volume')

                for pm in payment_modes:
                    ws.cell(row=row_num, column=1, value=pm['payment_mode'] or 'Unknown')
                    ws.cell(row=row_num, column=2, value=pm['total'])
                    ws.cell(row=row_num, column=3, value=pm['successful'])
                    ws.cell(row=row_num, column=4, value=pm['failed'])
                    ws.cell(row=row_num, column=5, value=round((pm['successful']/pm['total']*100) if pm['total'] > 0 else 0, 2))
                    ws.cell(row=row_num, column=6, value=pm['volume'] or 0)
                    ws.cell(row=row_num, column=7, value=pm['avg_amount'] or 0)
                    row_num += 1

            # Save file
            os.makedirs('/tmp/reports', exist_ok=True)
            filename = f"{report_type}_analytics_{user_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join('/tmp/reports', filename)
            wb.save(filepath)

            logger.info(f"Analytics Excel generated: {filepath}")

    except Exception as e:
        logger.error(f"Error generating analytics report: {str(e)}")

    return MockTask()


export_analytics_report.delay = export_analytics_report
